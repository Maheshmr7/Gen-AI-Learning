from playwright.sync_api import sync_playwright
from openpyxl import load_workbook
import time

# =====================================
# STEP 1 - READ EXCEL FILE
# =====================================

workbook = load_workbook("contacts.xlsx")
sheet = workbook.active

print("Reading Contacts...")

# =====================================
# STEP 2 - OPEN WHATSAPP WEB
# =====================================

with sync_playwright() as p:

    browser = p.chromium.launch(
        headless=False
    )

    page = browser.new_page()

    page.goto("https://web.whatsapp.com")

    input("Login to WhatsApp and press Enter...")

    # =====================================
    # STEP 3 - LOOP THROUGH CONTACTS
    # =====================================

    for row in sheet.iter_rows(min_row=2, values_only=True):

        name = row[0]
        phone = str(row[1])
        message = row[2]

        personalized_message = message.replace(
            "{name}",
            name
        )

        print(f"Sending message to {name}")

        # =====================================
        # STEP 4 - OPEN CHAT
        # =====================================

        page.goto(
            f"https://web.whatsapp.com/send?phone={phone}"
        )

        time.sleep(10)

        # =====================================
        # STEP 5 - TYPE MESSAGE
        # =====================================

        page.locator(
            '[contenteditable="true"]'
        ).last.click()

        page.keyboard.type(
            personalized_message
        )

        # =====================================
        # STEP 6 - SEND MESSAGE
        # =====================================

        page.keyboard.press("Enter")

        print("Message Sent")

        # =====================================
        # STEP 7 - TAKE SCREENSHOT
        # =====================================

        time.sleep(3)

        page.screenshot(
            path=f"screenshot_{name}.png"
        )

        print(
            f"Screenshot Saved for {name}"
        )

        time.sleep(5)

    print("All Messages Completed")

    input("Press Enter to Close Browser...")

    browser.close()
